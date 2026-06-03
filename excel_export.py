"""
excel_export.py
---------------
Export test session(s) from SQLite to a Juniper-branded Excel workbook.

Each session gets its own sheet. A summary sheet lists all sessions.
The workbook is intended for SharePoint sync.

Brand palette mirrors juniper-brand.css light-theme tokens:
  Navy    #1a1a2e  — brand bar / main headers
  Primary #1565c0  — sub-headers / accents
  Accent  #6e8cff  — highlight trim
  Offwhite #e8e7e3 — logo plate background
  Pass    #2e7d32 on #e8f5e9
  Fail    #b71c1c on #ffebee
  Muted   #f5f5f5 — alternating rows / meta labels
"""

import os
import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

from database import get_sessions, get_session, get_steps, get_thermal_tests, DB_PATH

# ── Juniper brand palette (ARGB hex, openpyxl format) ─────────────────────────
J_NAVY        = "FF1A1A2E"   # brand bar / primary headers
J_PRIMARY     = "FF1565C0"   # sub-headers
J_ACCENT      = "FF6E8CFF"   # trim / borders
J_OFFWHITE    = "FFE8E7E3"   # light plate / alternating rows
J_CHARCOAL    = "FF3C3C3D"   # secondary text
J_WHITE       = "FFFFFFFF"

# Data-row semantic colours
PASS_BG       = "FFE8F5E9"   # light green
PASS_FG       = "FF2E7D32"
FAIL_BG       = "FFFFEBEE"   # light red
FAIL_FG       = "FFB71C1C"
INCOMPLETE_BG = "FFFFF8E1"   # light amber
INCOMPLETE_FG = "FFF57F17"

# Meta label column
META_BG       = "FFF0F2F5"   # --bg equivalent

# ── Font name: Calibri is the closest Excel built-in to Poppins geometry ──────
BRAND_FONT = "Calibri"

STATUS_FLAGS = {
    1:     "Internal fault",
    2:     "Over voltage",
    4:     "Line too low",
    8:     "DUT breakdown",
    16:    "HOLD timeout",
    32:    "User aborted",
    64:    "GB over-compliance",
    128:   "Arc detected",
    256:   "Below min limit",
    512:   "Above max limit",
    1024:  "IR steady current fail",
    2048:  "Interlock failure",
    4096:  "Switch matrix error",
    8192:  "Overheated",
    16384: "Cannot control output",
    32768: "GB wiring error",
    65536: "Drive instability",
}


# ── Helpers ────────────────────────────────────────────────────────────────────

def _border(color="FFD0D7E3", style="thin"):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def _accent_border():
    """Left-only accent bar in J_ACCENT for meta label column."""
    return Border(left=Side(style="medium", color=J_ACCENT))

def _navy_cell(ws, row, col, value, size=11, bold=True, span_end_col=None):
    """Dark navy header cell — used for sheet title banners."""
    if span_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=span_end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=BRAND_FONT, bold=bold, size=size, color=J_WHITE)
    c.fill      = PatternFill("solid", fgColor=J_NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    return c

def _primary_cell(ws, row, col, value, size=9, bold=True, wrap=False, span_end_col=None):
    """Mid-blue sub-header cell."""
    if span_end_col:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=span_end_col)
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=BRAND_FONT, bold=bold, size=size, color=J_WHITE)
    c.fill      = PatternFill("solid", fgColor=J_PRIMARY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c

def _meta_label(ws, row, col, value):
    """Label cell for the session meta block (left column)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=BRAND_FONT, bold=True, size=9, color=J_CHARCOAL)
    c.fill      = PatternFill("solid", fgColor=META_BG)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border    = _accent_border()
    return c

def _meta_value(ws, row, col, value):
    """Value cell for the session meta block (right column)."""
    c = ws.cell(row=row, column=col, value=str(value) if value is not None else "—")
    c.font      = Font(name=BRAND_FONT, size=9, color=J_CHARCOAL)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=2)
    c.border    = _border(color="FFE0E0E0")
    return c

def _data_cell(ws, row, col, value, number_format=None, bold=False,
               bg=None, fg=J_CHARCOAL, size=9):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name=BRAND_FONT, bold=bold, size=size, color=fg)
    c.alignment = Alignment(vertical="center", indent=1)
    c.border    = _border()
    if number_format:
        c.number_format = number_format
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    return c

def decode_flags(flags: int) -> str:
    if flags == 0:
        return "PASS"
    msgs = [desc for bit, desc in STATUS_FLAGS.items() if flags & bit]
    return "; ".join(msgs) if msgs else f"Unknown (0x{flags:X})"


# ── Branded title banner (rows 1–2) ───────────────────────────────────────────

def _write_title_banner(ws, num_cols: int, subtitle: str) -> int:
    """
    Write a two-row Juniper brand header spanning all columns.
    Returns the next available row number.
    """
    # Row 1: navy brand bar — "JUNIPER DESIGN  ·  V71 HiPot Controller"
    c1 = _navy_cell(ws, 1, 1, "JUNIPER DESIGN  ·  V71 HiPot Controller",
                    size=13, span_end_col=num_cols)
    ws.row_dimensions[1].height = 30

    # Right-align a "Generated" timestamp in the same merged cell's last physical cell
    ts_cell = ws.cell(row=1, column=num_cols,
                      value=f"Generated {datetime.datetime.now():%Y-%m-%d %H:%M}")
    ts_cell.font      = Font(name=BRAND_FONT, size=8, color="FFAAAACC", italic=True)
    ts_cell.fill      = PatternFill("solid", fgColor=J_NAVY)
    ts_cell.alignment = Alignment(horizontal="right", vertical="center", indent=2)

    # Row 2: accent-coloured subtitle bar
    c2 = ws.cell(row=2, column=1, value=subtitle)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    c2.font      = Font(name=BRAND_FONT, bold=True, size=10, color=J_WHITE)
    c2.fill      = PatternFill("solid", fgColor=J_PRIMARY)
    c2.alignment = Alignment(horizontal="left", vertical="center", indent=3)
    ws.row_dimensions[2].height = 20

    return 3  # next free row


# ── Session detail sheet ───────────────────────────────────────────────────────

def _write_session_sheet(wb: openpyxl.Workbook, session: dict, steps: list[dict]) -> None:
    title = f"S{session['id']}-{(session.get('serial_number') or 'DUT')[:18]}"
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False

    # Column widths
    col_widths = [22, 32, 14, 14, 16, 16, 14, 34]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    NUM_COLS = len(col_widths)

    # Title banner
    next_row = _write_title_banner(
        ws, NUM_COLS,
        subtitle=f"Test Report  ·  Session #{session['id']}  ·  "
                 f"DUT: {session.get('serial_number') or '—'}  ·  "
                 f"Part: {session.get('part_number') or '—'}"
    )

    # ── Meta block ──────────────────────────────────────────────────────────
    meta = [
        ("Session ID",    session["id"]),
        ("Started",       session.get("started_at", "")),
        ("Finished",      session.get("finished_at") or "—"),
        ("Operator",      session.get("operator") or "—"),
        ("Part Number",   session.get("part_number") or "—"),
        ("DUT Serial",    session.get("serial_number") or "—"),
        ("Device Model",  session.get("device_model") or "—"),
        ("Device Serial", session.get("device_serial") or "—"),
        ("Firmware",      session.get("firmware") or "—"),
        ("Notes",         session.get("notes") or "—"),
    ]
    for label, value in meta:
        _meta_label(ws, next_row, 1, label)
        _meta_value(ws, next_row, 2, value)
        # Shade remaining columns in the meta rows for a clean look
        for col in range(3, NUM_COLS + 1):
            c = ws.cell(row=next_row, column=col)
            c.fill = PatternFill("solid", fgColor=J_OFFWHITE)
        ws.row_dimensions[next_row].height = 16
        next_row += 1

    # ── Overall result banner ───────────────────────────────────────────────
    overall = session.get("overall_result")
    passed  = session.get("passed")
    next_row += 1   # blank spacer
    ws.merge_cells(start_row=next_row, start_column=1,
                   end_row=next_row, end_column=NUM_COLS)
    rc = ws.cell(row=next_row, column=1)
    if passed == 1:
        rc.value = "✓   OVERALL RESULT:   PASS"
        rc.font  = Font(name=BRAND_FONT, bold=True, size=13, color=J_WHITE)
        rc.fill  = PatternFill("solid", fgColor="FF2E7D32")
    elif passed == 0:
        rc.value = f"✗   OVERALL RESULT:   FAIL  —  {decode_flags(overall or 0)}"
        rc.font  = Font(name=BRAND_FONT, bold=True, size=13, color=J_WHITE)
        rc.fill  = PatternFill("solid", fgColor="FFB71C1C")
    else:
        rc.value = "—   INCOMPLETE / IN PROGRESS"
        rc.font  = Font(name=BRAND_FONT, bold=True, size=13, color=INCOMPLETE_FG)
        rc.fill  = PatternFill("solid", fgColor=INCOMPLETE_BG)
    rc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 26
    next_row += 2   # spacer before steps table

    # ── Steps table ─────────────────────────────────────────────────────────
    step_headers = ["Step #", "Type", "Phase", "Elapsed (s)",
                    "Level (V/A)", "Leakage / R", "Arc (A)", "Status / Flags"]
    for col, h in enumerate(step_headers, 1):
        _primary_cell(ws, next_row, col, h, wrap=True)
    ws.row_dimensions[next_row].height = 32
    next_row += 1

    for step in steps:
        sf    = step.get("status_flags", 0) or 0
        p     = step.get("passed")
        bg    = PASS_BG if p == 1 else (FAIL_BG if p == 0 else None)
        fg    = PASS_FG if p == 1 else (FAIL_FG if p == 0 else J_CHARCOAL)

        _data_cell(ws, next_row, 1, step.get("step_number"),   bg=bg, bold=True, fg=fg)
        _data_cell(ws, next_row, 2, step.get("step_type", ""), bg=bg, bold=True, fg=fg)
        _data_cell(ws, next_row, 3, step.get("phase", ""),     bg=bg, fg=fg)
        _data_cell(ws, next_row, 4, step.get("elapsed_s"),     bg=bg, fg=fg, number_format="0.000")
        _data_cell(ws, next_row, 5, step.get("level"),         bg=bg, fg=fg, number_format="0.000E+00")
        _data_cell(ws, next_row, 6, step.get("measurement"),   bg=bg, fg=fg, number_format="0.000E+00")
        _data_cell(ws, next_row, 7, step.get("arc_a"),         bg=bg, fg=fg, number_format="0.000E+00")
        _data_cell(ws, next_row, 8, decode_flags(sf),          bg=bg, fg=fg, bold=(sf != 0))
        ws.row_dimensions[next_row].height = 15
        next_row += 1

    # ── Juniper footer rule ─────────────────────────────────────────────────
    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1,
                   end_row=next_row, end_column=NUM_COLS)
    ft = ws.cell(row=next_row, column=1,
                 value="Designed and built by Juniper Design  ·  juniperdesign.com")
    ft.font      = Font(name=BRAND_FONT, size=8, italic=True, color="FF9090A8")
    ft.fill      = PatternFill("solid", fgColor=J_NAVY)
    ft.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 18


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _write_summary_sheet(wb: openpyxl.Workbook, sessions: list[dict]) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    col_widths = [10, 22, 22, 20, 20, 10, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    NUM_COLS = len(col_widths)

    # Stats for subtitle
    total  = len(sessions)
    passed = sum(1 for s in sessions if s.get("passed") == 1)
    failed = total - passed
    next_row = _write_title_banner(
        ws, NUM_COLS,
        subtitle=f"All Test Sessions  ·  {total} total  ·  {passed} passed  ·  {failed} failed"
    )

    # Column headers
    hdrs = ["Session #", "Started", "Finished", "Part Number", "DUT Serial", "Result", "Steps"]
    for col, h in enumerate(hdrs, 1):
        _primary_cell(ws, next_row, col, h)
    ws.row_dimensions[next_row].height = 20
    next_row += 1

    for s in sessions:
        p = s.get("passed")
        bg         = PASS_BG if p == 1 else (FAIL_BG if p == 0 else None)
        fg         = PASS_FG if p == 1 else (FAIL_FG if p == 0 else J_CHARCOAL)
        result_str = "PASS"    if p == 1 else ("FAIL" if p == 0 else "—")

        _data_cell(ws, next_row, 1, s.get("id"),                     bg=bg, fg=fg, bold=True)
        _data_cell(ws, next_row, 2, s.get("started_at", ""),         bg=bg, fg=fg)
        _data_cell(ws, next_row, 3, s.get("finished_at") or "—",     bg=bg, fg=fg)
        _data_cell(ws, next_row, 4, s.get("part_number") or "—",     bg=bg, fg=fg)
        _data_cell(ws, next_row, 5, s.get("serial_number") or "—",   bg=bg, fg=fg)
        _data_cell(ws, next_row, 6, result_str, bg=bg, fg=fg, bold=True)
        _data_cell(ws, next_row, 7, None, bg=bg, fg=fg)
        ws.row_dimensions[next_row].height = 15
        next_row += 1

    # Auto-filter on the header row (row 3 = next_row - len(sessions) - 1)
    hdr_row = 3
    ws.auto_filter.ref = (
        f"A{hdr_row}:{get_column_letter(NUM_COLS)}{hdr_row + len(sessions)}"
    )

    # Footer
    next_row += 1
    ws.merge_cells(start_row=next_row, start_column=1,
                   end_row=next_row, end_column=NUM_COLS)
    ft = ws.cell(row=next_row, column=1,
                 value="Designed and built by Juniper Design  ·  juniperdesign.com")
    ft.font      = Font(name=BRAND_FONT, size=8, italic=True, color="FF9090A8")
    ft.fill      = PatternFill("solid", fgColor=J_NAVY)
    ft.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[next_row].height = 18


# ── Public API ────────────────────────────────────────────────────────────────

def _write_pec0063_sheet(wb: openpyxl.Workbook, rows: list[dict]) -> None:
    """Write a PEC-0063 thermal qualification results sheet."""
    ws = wb.create_sheet("PEC-0063 Thermal Results")
    ws.sheet_view.showGridLines = False

    # ── Banner ────────────────────────────────────────────────────────────────
    _navy_cell(ws, 1, 1, "PEC-0063 Thermal Qualification — Test Results", size=13, span_end_col=10)
    ws.row_dimensions[1].height = 22
    _primary_cell(ws, 2, 1,
                  "65 W USB-C Power Supply (20 V / 3.25 A) · 66 W test load · Source: DSR 10/7/2025",
                  span_end_col=10)
    ws.row_dimensions[2].height = 16

    # ── UL limits reference block ─────────────────────────────────────────────
    ws.cell(row=3, column=1, value="UL Standard Reference").font = Font(
        name=BRAND_FONT, bold=True, size=8, color=J_CHARCOAL)
    limits = [
        ("UL 962A", "Nonmetallic surface: ≤ 95 °C absolute"),
        ("UL 1310", "Nonmetallic surface: ≤ +50 °C above ambient  |  Metallic: ≤ +30 °C above ambient"),
    ]
    for offset, (std, desc) in enumerate(limits):
        r = 4 + offset
        ws.cell(row=r, column=1, value=std).font = Font(name=BRAND_FONT, bold=True, size=8, color=J_PRIMARY)
        ws.cell(row=r, column=2, value=desc).font = Font(name=BRAND_FONT, size=8, color=J_CHARCOAL)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)

    # ── Column headers (row 7) ────────────────────────────────────────────────
    headers = [
        "Date", "Housing", "Standard", "Surface",
        "Load (W)", "Tcase (°C)", "Ambient (°C)", "Rise ΔT (°C)",
        "Limit (°C)", "Margin (°C)", "Result", "Note",
    ]
    HR = 7
    for col, h in enumerate(headers, 1):
        _primary_cell(ws, HR, col, h, size=9, bold=True)
    ws.row_dimensions[HR].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    RESULT_COLORS = {
        "PASS":     (PASS_BG,       PASS_FG),
        "MARGINAL": (INCOMPLETE_BG, INCOMPLETE_FG),
        "FAIL":     (FAIL_BG,       FAIL_FG),
    }

    for dr, row in enumerate(rows):
        r = HR + 1 + dr
        result = (row.get("result") or "").upper()
        bg, fg = RESULT_COLORS.get(result, (J_OFFWHITE, J_CHARCOAL))
        alt_bg = "FFF8F9FA" if dr % 2 else "FFFFFFFF"

        def _cell(col, val, bold=False, align="left", num_format=None, color=None):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name=BRAND_FONT, size=9, bold=bold,
                               color=(fg if col == len(headers) else J_CHARCOAL))
            c.fill      = PatternFill("solid", fgColor=(bg if col == len(headers) else alt_bg))
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border    = _border()
            if num_format:
                c.number_format = num_format
            return c

        _cell(1,  row.get("started_at", "")[:10])
        _cell(2,  row.get("housing_key", ""))
        _cell(3,  row.get("standard", ""))
        _cell(4,  row.get("surface_type", ""))
        _cell(5,  row.get("dc_load_w"),   align="center", num_format="0.0")
        _cell(6,  row.get("tcase_c"),     align="center", num_format="0.0")
        _cell(7,  row.get("ambient_c"),   align="center", num_format="0.0")
        _cell(8,  row.get("rise_c"),      align="center", num_format="+0.0;-0.0")
        _cell(9,  row.get("limit_c"),     align="center", num_format="0")
        _cell(10, row.get("margin_c"),    align="center", num_format="+0.0;-0.0")
        # Result cell gets coloured bg
        rc = ws.cell(row=r, column=11, value=result)
        rc.font      = Font(name=BRAND_FONT, size=9, bold=True, color=fg)
        rc.fill      = PatternFill("solid", fgColor=bg)
        rc.alignment = Alignment(horizontal="center", vertical="center")
        rc.border    = _border()
        _cell(12, row.get("note", ""))

    # ── Baseline reference block (from DSR) ───────────────────────────────────
    if not rows:
        # If no new test data, still show the baselines
        baseline = [
            ("2025-10-20", "UDM_Single", "UL_1310", "nonmetallic", 66, 93, 23, 70, 50, -20, "FAIL", "Test stopped after 2 hrs"),
            ("2025-10-20", "DSK_Single", "UL_1310", "nonmetallic", 66, 85, 22, 63, 50, -13, "FAIL", ""),
            ("2025-10-20", "NCP_Single", "UL_1310", "nonmetallic", 66, 77, 24, 53, 50, -3,  "FAIL", ""),
            ("2025-10-20", "DSK_Triple", "UL_1310", "nonmetallic", 66, 75, 24, 51, 50, -1,  "FAIL", ""),
            ("2025-10-20", "UDM_Triple", "UL_1310", "nonmetallic", 66, 85, 24, 61, 50, -11, "FAIL", ""),
            ("2025-10-20", "DSK_Double", "UL_1310", "nonmetallic", 66, 85, 24, 61, 50, -11, "FAIL", ""),
        ]
        note_row = HR + 1
        _navy_cell(ws, note_row, 1, "Baseline results from DSR 10/7/2025 (no new station data yet)",
                   size=9, bold=False, span_end_col=12)
        for dr, b in enumerate(baseline):
            r = note_row + 1 + dr
            vals = list(b)
            result = vals[10].upper()
            bg, fg = RESULT_COLORS.get(result, (J_OFFWHITE, J_CHARCOAL))
            alt_bg = "FFF8F9FA" if dr % 2 else "FFFFFFFF"
            for col, val in enumerate(vals, 1):
                c = ws.cell(row=r, column=col, value=val)
                c.font      = Font(name=BRAND_FONT, size=9, italic=True,
                                   color=(fg if col == 11 else J_CHARCOAL))
                c.fill      = PatternFill("solid", fgColor=(bg if col == 11 else alt_bg))
                c.alignment = Alignment(horizontal="center" if col > 4 else "left", vertical="center")
                c.border    = _border()

    # ── Column widths ─────────────────────────────────────────────────────────
    widths = [12, 14, 10, 12, 9, 12, 13, 13, 10, 12, 10, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def export_to_excel(output_path: str,
                    session_ids: Optional[list[int]] = None,
                    db_path: str = DB_PATH) -> str:
    """
    Export sessions and PEC-0063 thermal results to a Juniper-branded Excel file.
    If session_ids is None, exports all sessions (most recent first, up to 500).
    Returns the path to the written file.
    """
    if session_ids:
        sessions = [get_session(sid, db_path) for sid in session_ids]
        sessions = [s for s in sessions if s]
    else:
        sessions = get_sessions(limit=500, db_path=db_path)

    wb = openpyxl.Workbook()
    _write_summary_sheet(wb, sessions)

    for session in sessions:
        steps = get_steps(session["id"], db_path)
        _write_session_sheet(wb, session, steps)

    # PEC-0063 thermal qualification sheet
    thermal_rows = get_thermal_tests(limit=500, db_path=db_path)
    _write_pec0063_sheet(wb, thermal_rows)

    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import sys
    out = sys.argv[1] if len(sys.argv) > 1 else "hipot_results.xlsx"
    path = export_to_excel(out)
    print(f"Exported to {path}")
